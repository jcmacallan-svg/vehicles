#!/usr/bin/env python3
"""
Apply classifications from an Excel to app/data.json.

Expected Excel headers (row 1):
- asset
- answer
- class

Run from repo root, e.g.:
  python3 tools/apply_classifications.py --xlsx spreadsheets/vehicle_classification_template_filled.xlsx --data app/data.json
"""
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

def load_xlsx(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h:i for i,h in enumerate(headers)}
    for h in ("asset","answer","class"):
        if h not in idx:
            raise SystemExit(f"Excel missing header '{h}'. Found: {headers}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        asset = (r[idx["asset"]] or "").strip()
        if not asset:
            continue
        answer = (r[idx["answer"]] or "").strip()
        klass = (r[idx["class"]] or "").strip()
        rows.append((asset, answer, klass))
    return rows

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--data", required=True)
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    data_path = Path(args.data)

    data = json.loads(data_path.read_text(encoding="utf-8"))
    mapping = {a: (ans, k) for a, ans, k in load_xlsx(xlsx)}

    updated = 0
    for q in data.get("questions", []):
        a = q.get("asset","")
        if a in mapping:
            ans, k = mapping[a]
            if ans: q["answer"] = ans
            if k: q["class"] = k
            updated += 1

    data_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"✅ Updated {updated} questions in {data_path}")

if __name__ == "__main__":
    main()
